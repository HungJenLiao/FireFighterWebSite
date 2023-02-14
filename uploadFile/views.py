from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
from .models import Emergency
from .forms import EmergencyForm
from django.contrib.auth.models import User, auth
from django .contrib.auth.decorators import login_required

from django.forms.models import model_to_dict
from django.http import JsonResponse


from os import mkdir
from os.path import isdir, abspath, dirname, join
from openpyxl import load_workbook

BASE_DIR = dirname(dirname(abspath(__file__)))


# Create your views here.
def login(request):
    if request.user.is_authenticated:
        return redirect('/dashboard/')
    username = request.POST.get('username')
    password = request.POST.get('password')
    user = auth.authenticate(username = username, password = password)
    if user is not None and user.is_active:
        auth.login(request, user)
        return redirect('/dashboard/')
    else:
        return render(request, 'login.html')
        
def logout(request):
    auth.logout(request)
    return redirect('/accounts/login/')

@login_required
def main_template(request):
    #Login FullName
    fullname = request.user.get_full_name()
    context = {'fullname': fullname}
    return render(request, 'main_template.html', context)

@login_required
def uploadFile(request):
    #get user name
    fullname = request.user.get_full_name()

    if request.method == 'POST':
        #創建資料夾
        uploadDir = BASE_DIR + '/upload'
        if not isdir(uploadDir):
            mkdir(uploadDir)
        print(uploadDir)
        #抓取上傳資料
        uploadedFile = request.FILES.get('uploadFile')
        if not uploadedFile:
            return render(request, 'uploadFile.html', {'msg': '沒有選擇文件'})
        if not uploadedFile.name.endswith('.xlsx'):
            return render(request, 'uploadFile.html', {'msg': '必須選擇xlsx文件'})
        # #上傳資料
        # with open(uploadedFile.name, 'wb') as fp:
        #     for chunk in uploadedFile.chunks():
        #         fp.write(chunk)
        # #導入數據庫
        # ws = load_workbook(uploadedFile).worksheets[0]
        # #刪去欄位名稱
        # ws.delete_rows(0, 1)
        # max_column = ws.max_column
        # #刪除所有資料
        # Emergency.objects.all().delete()
        # for index, row in enumerate(ws.rows):
        #     #填補缺失值
        #     for num in range(0, max_column):
        #         if row[num].value == None:
        #             row[num].value = 0
        #     Emergency.objects.create(time = row[0].value, 
        #                              unit = row[1].value,
        #                              category = row[2].value, 
        #                              detail = row[3].value, 
        #                              location = row[4].value)
        nameTransform(uploadedFile)
        return render(request, 'uploadFile.html', {'fullname': fullname})
    return render(request, 'uploadFile.html', {'fullname': fullname})

def nameTransform(fp):
    #導入上傳的資料
    import pandas as pd
    df = pd.read_excel(fp)
    #######################################################
    #導入 search.py 引用 data_cleaning 去除缺失值以及更改時間格式,
    #search_addr 利用 selenium 上國土測繪網站進行爬蟲
    #search.write_to_file("XXXXX.xlsx", df)儲存到local進行測試
    #######################################################
    import search
    df_info = search.data_cleaning(df)
    df_locations = df_info[0].iloc[117:121]
    # df_locations_num = df_info[1]
    df_locations_num = 3
    for num in range(117, 121):
        df_locations[num] = search.search_addr(df_locations[num])
        print(num)
    df["發生地點"] = df_locations
    
    search.write_to_file("測試.xlsx", df)

@login_required
def dashboard(request):
    #get username
    fullname = request.user.get_full_name()
    #get total cases
    emergency_count = Emergency.objects.count()
    #get barChartInfo
    emergency_objs = Emergency.objects.all().order_by("id")
    loc_list = []
    for obj in emergency_objs:
        if obj:
            loc_dict = model_to_dict(obj, fields = ["location"])
            loc_list.append(loc_dict["location"])
    response = barChartInfo(loc_list)

    return render(request, 'dashboard_index.html', {'emergency_count': emergency_count, 
                                                    'fullname': fullname, 
                                                    'response': response})

def dashboard_model(request):
    return render(request, 'dashboard_model.html')

@login_required
def emergency_list(request):
    #Login FullName
    fullname = request.user.get_full_name()
    #Get Emergency List Inofrmation
    emergency_list = Emergency.objects.all().order_by("id")
    context = {'fullname': fullname, 'emergency_list': emergency_list}
    return render(request, 'emergency_list.html', context)

@login_required
def emergency_list_update(request, Em_id):
    #Login FullName
    fullname = request.user.get_full_name()
    #Get Each ID
    emergency_obj = Emergency.objects.get(pk = Em_id)
    #Form
    form = EmergencyForm
    context = {'fullname': fullname, 'emergency_obj': emergency_obj, 'form': form}
    return render(request, 'emergency_list_update.html', context)

def emergency_list_edit(request):
    #Login FullName
    fullname = request.user.get_full_name()
    #Form
    submitted = False
    if request.method == "POST":
        form = EmergencyForm(request.POST)
        if form.is_valid():
            form.save()
            context = {'fullname': fullname, 'form': form}
            return HttpResponseRedirect('/emergency_list/edit?submitted=True/', context)
    else:
        form = EmergencyForm
        if 'submitted' in request.GET:
            submitted = True
    context = {'fullname': fullname, 'form': form, 'submitted': submitted}
    return render(request, 'emergency_list_edit.html', context)

@login_required
def member_list(request):
    #Login FullName
    fullname = request.user.get_full_name()
    #Get user information from admin model
    all_users = User.objects.values()

    context = {'fullname': fullname, 'all_users': all_users}
    return render(request, 'member_list.html', context)

@login_required
def logrecord_list(request):

    if request.user.is_authenticated:
        Username = request.user.get_username()
        Fullname = request.user.get_full_name()
        Datejoined = request.user.date_joined
        Lastlogin = request.user.last_login
    return render(request, 'logrecord_list.html', {'Username': Username, 
                                                    'Fullname': Fullname, 
                                                    'Datejoined': Datejoined, 
                                                    'Lastlogin': Lastlogin})
                                                    
@login_required
def news(request):
    #Login FullName
    fullname = request.user.get_full_name()
    context = {'fullname': fullname}
    return render(request, 'news.html', context)

def barChartInfo(file):
    #Get Data From SQL
    emergency_obj = Emergency.objects.all().order_by("id")
    import counting
    response = counting.count_num(file)
    return response

def test(request):
    # emergency_objs = Emergency.objects.all().order_by("id")
    # if emergency_objs:
    #     locs = emergency_objs.values("location")
    #     loc_list = [loc["location"] for loc in locs]

    emergency_objs = Emergency.objects.all().order_by("id")
    loc_list = []
    for obj in emergency_objs:
        if obj:
            loc_dict = model_to_dict(obj, fields = ["location"])
            loc_list.append(loc_dict["location"])

    # print(loc_list)
    import counting
    response = counting.count(loc_list)
    context = {"response": response}
    # return render(request, 'test.html', context)
    return render(request, 'test.html', context) 

